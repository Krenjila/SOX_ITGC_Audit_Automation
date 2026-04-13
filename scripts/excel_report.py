# excel_report.py
import pandas as pd
import sqlite3
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

SOD_COLOR    = "AEC6CF"
TERM_COLOR   = "FDFD96"
GHOST_COLOR  = "C1E1C1"
HEADER_COLOR = "2F4F8F"
TITLE_COLOR  = "1C1C3A"

def load_findings(csv_path):
    df = pd.read_csv(csv_path, parse_dates=["termination_date", "last_login", "hire_date"])
    for col in ["termination_date", "last_login", "hire_date"]:
        df[col] = df[col].astype(str).replace("NaT", None)
    conn = sqlite3.connect(":memory:")
    df.to_sql("users", conn, if_exists="replace", index=False)
    findings = {
        "SOD Violations": pd.read_sql("SELECT user_id, name, department, status, termination_date, last_login, role, access_level FROM users WHERE role = 'Initiator+Approver'", conn),
        "Terminated Still Active": pd.read_sql("SELECT user_id, name, department, status, termination_date, last_login, role, access_level FROM users WHERE status = 'Terminated' AND last_login > termination_date AND termination_date >= date('now', '-730 days')", conn),
        "Ghost Accounts": pd.read_sql("SELECT user_id, name, department, status, termination_date, last_login, role, access_level FROM users WHERE status = 'Terminated' AND termination_date < date('now', '-730 days')", conn),
    }
    conn.close()
    return findings

def build_sheet(ws, df, fill_color, sheet_title, finding_type):
    fill        = PatternFill("solid", fgColor=fill_color)
    header_fill = PatternFill("solid", fgColor=HEADER_COLOR)
    title_fill  = PatternFill("solid", fgColor=TITLE_COLOR)
    thin        = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    columns     = ["user_id", "name", "department", "status", "termination_date", "last_login", "role", "access_level"]
    col_widths  = [10, 22, 16, 14, 18, 18, 20, 14]
    ws.merge_cells(f"A1:{get_column_letter(len(columns))}1")
    ws["A1"] = f"SOX ITGC Audit Workpaper — {sheet_title}"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = title_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells(f"A2:{get_column_letter(len(columns))}2")
    ws["A2"] = f"Generated: {date.today().strftime('%B %d, %Y')}     |     Finding Type: {finding_type}     |     Total Findings: {len(df)}"
    ws["A2"].font = Font(italic=True, size=11, color="555555")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.append([])
    headers = [c.replace("_", " ").title() for c in columns]
    ws.append(headers)
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col_num)
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin
    ws.row_dimensions[4].height = 18
    for _, row in df.iterrows():
        ws.append([row[c] for c in columns])
        for col_num in range(1, len(columns) + 1):
            cell = ws.cell(row=ws.max_row, column=col_num)
            cell.fill = fill
            cell.border = thin
            cell.alignment = Alignment(horizontal="center")
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width
    ws.freeze_panes = "A5"

def build_report(csv_path, output_path):
    findings = load_findings(csv_path)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_config = [
        ("SOD Violations",          SOD_COLOR,   "Segregation of Duties Violation"),
        ("Terminated Still Active", TERM_COLOR,  "Terminated User Still Accessing System"),
        ("Ghost Accounts",          GHOST_COLOR, "Ghost Account — Inactive 2+ Years"),
    ]
    for sheet_name, color, finding_type in sheet_config:
        ws = wb.create_sheet(title=sheet_name)
        build_sheet(ws, findings[sheet_name], color, sheet_name, finding_type)
        print(f"{sheet_name}: {len(findings[sheet_name])} findings")
    wb.save(output_path)
    total = sum(len(v) for v in findings.values())
    print(f"\nReport saved → {output_path}")
    print(f"Total findings: {total}")

if __name__ == "__main__":
    build_report("../data/users.csv", "../data/sox_audit_workpaper.xlsx")
